import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1): #first argument indicates starting point of row
    cell = sheet.cell(row, 3) #The second argument identifies number of the column
    corrected_price = cell.value * 0.9
    corrected_cell = sheet.cell(row, 4)
    corrected_cell.value = corrected_price
    sheet.cell(1,4).value = 'Correct price'
    sheet.cell(4, 1).value = 1003


for row in range (6, 15):
    cell = sheet.cell(row, 1)
    cell.value = row

for column in range (1, 15): #In this case the 1, is the starting point because the loop now will run horizontally
    cell = sheet.cell(16, column)
    cell.value = column


for row in range (6, 15):
    cell = sheet.cell(row, 2)
    cell.value = f"David {row}"


values = Reference(sheet, 
    min_row=2,
    max_row=5,
    min_col=4,
    max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')


wb.save('transactions2.xlsx')